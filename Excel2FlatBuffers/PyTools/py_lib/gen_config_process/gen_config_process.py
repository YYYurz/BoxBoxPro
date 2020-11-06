import os
import re
import openpyxl
import importlib
import importlib.util
import collections
import inspect
import sys
import stringcase
from openpyxl.reader import excel
from py_lib import flatbuffers
from py_lib import util
from . import build_seq
from . import build_struct


def flatc(dir_map, flatc_exec, file_path):
    cmd_list = [flatc_exec, "-o {}".format(dir_map.out_py_dir), "-p", file_path]
    cmd = " ".join(cmd_list)
    util.run_shell(cmd, wait=True, output=False)

    cmd_list = [flatc_exec, "-o {}".format(dir_map.out_cs_dir), "-n --gen-onefile", file_path]
    cmd = " ".join(cmd_list)
    util.run_shell(cmd, wait=True, output=False)

    cmd_list = [flatc_exec, "-o {}".format(dir_map.out_lua_dir), "-l", file_path]
    cmd = " ".join(cmd_list)
    util.run_shell(cmd, wait=True, output=False)

    cmd_list = [flatc_exec, "-o {}".format(dir_map.out_cpp_dir), "-c", file_path]
    cmd = " ".join(cmd_list)
    util.run_shell(cmd, wait=True, output=False)

    name_list = list()
    with open(file_path) as f:
        for line in f:
            pos = line.find("//")
            if pos > -1:
                line = line[:pos]
            r = re.search(r"\s*(\w+)\s*:(\[?)(\w+)(\]?)\s*;", line)
            if r:
                name_list.append(r.groups())
    return name_list


def read_proto(dir_map, key):
    rel_path = os.path.relpath(dir_map.out_py_dir, dir_map.project_dir).replace("\\", "/")
    lib_path = rel_path.replace("/", ".")
    rel_path = os.path.relpath(dir_map.out_py_dir, os.getcwd()).replace("\\", "/")
    key = stringcase.capitalcase(key)

    sys.path.append(os.path.abspath(util.get_current_path()))
    lib_util = importlib.util

    spec = lib_util.spec_from_file_location("{}.{}.{}".format(lib_path, dir_map.fbs_namespace, key),
                                            "{}/{}/{}.py".format(rel_path, dir_map.fbs_namespace, key))
    module = lib_util.module_from_spec(spec)
    spec.loader.exec_module(module)

    spec = lib_util.spec_from_file_location("{}.{}.{}List".format(lib_path, dir_map.fbs_namespace, key),
                                            "{}/{}/{}List.py".format(rel_path, dir_map.fbs_namespace, key))
    list_module = lib_util.module_from_spec(spec)
    spec.loader.exec_module(list_module)

    sys.path.pop()
    return module, list_module


def gen_bytes(dir_map, modules, dict_parse, dont_skip):
    for dirpath, _, filenames in os.walk(dir_map.config_dir):
        filenames = filter(lambda x: re.match(r"^[^~]+\.xlsx$", x), filenames)
        for filename in filenames:
            fname, _ = os.path.splitext(filename)
            pos = fname.find("_")
            if pos >= 0:
                fname = fname[:pos]
            if fname not in dict_parse:
                print("{} ignoreed".format(fname))
                continue
            file_path = os.path.join(dirpath, filename)
            book = openpyxl.reader.excel.load_workbook(file_path, data_only=True)
            sheet = book[book.sheetnames[0]]
            can_gen, m_name, k_title = pre_gen(sheet.title, modules, dict_parse)
            if not can_gen:
                continue
            if sheet.max_row < 1 or sheet.max_column < 1:
                print("{} sheet is None".format(sheet.title))
                continue
            read_excel_and_gen(dir_map, filename, modules, dict_parse, sheet, m_name,
                               k_title, dont_skip)
            book.close()


def format_excel_value(cell, type_str, can_be_none=True):
    if cell.value is None:
        return format_none_value(type_str, can_be_none)
    cvalue = cell.value
    if type_str == "string":
        return str(cell.value)
    if isinstance(cvalue, str):
        cvalue = cvalue.strip()
        if cvalue == "":
            return format_none_value(type_str, can_be_none)
    if type_str in ("bool", "byte", "ubyte", "short", "ushort", "int", "uint"):
        return int(cvalue)
    if type_str in ("float", "FP"):
        return float(cvalue)
    return cvalue


def format_str_value(value: str, type_str: str, can_be_none=True):
    if value is None:
        return format_none_value(type_str, can_be_none)
    if type_str == "string":
        return value
    value = value.strip()
    if value == "":
        return format_none_value(type_str, can_be_none)
    if type_str in ("bool", "byte", "ubyte", "short", "ushort", "int", "uint"):
        return int(value)
    if type_str in ("float", "FP"):
        return float(value)
    return value


def format_none_value(type_str: str, can_be_none):
    if can_be_none:
        return None
    if type_str in ("bool", "byte", "ubyte", "short", "ushort", "int", "uint"):
        return int(0)
    if type_str in ("float", "FP"):
        return float(0)
    if type_str == "string":
        return str("0")
    raise Exception("Can't be none")


def format_type(type_str):
    ret = stringcase.capitalcase(type_str)
    if ret in ("Int", "Uint", "Float"):
        return ret + "32"
    if ret == "Ubyte":
        return "Uint8"
    if ret == "Short":
        return "Int16"
    if ret == "UShort":
        return "Uint16"
    if ret == "Long":
        return "Int64"
    if ret == "Ulong":
        return "Uint64"
    if ret == "Double":
        return "Float64"
    return ret


def trim_blank(cell_list):
    while len(cell_list):
        v = cell_list[-1]
        if v[0].value is not None:
            break
        cell_list.pop()


def store_list_func_and_arg(bs, mb, builder, temp_value, typ, modules):
    if typ == "FP":
        ml = modules["FP"].__dict__
        temp_value = int(temp_value * 4294967296)
        func_name = "Create{}".format(typ)
        bs.store_func(ml["Create{}".format(typ)], [func_name, builder, temp_value])
        return
    if typ == "string":
        temp_value = builder.CreateString(temp_value or "")
        func_name = "PrependUOffsetTRelative"
    elif typ == "FP":
        ml = modules["FP"].__dict__
        temp_value = build_struct.BuildStruct(ml["Create{}".format(typ)],
                                              [builder, int(temp_value * 4294967296)])
        func_name = "PrependUOffsetTRelative"
    else:
        func_name = "Prepend%s" % format_type(typ)
    bs.store_func(mb[func_name], [func_name, temp_value])


def read_excel_and_gen(dir_map, filename, modules, dict_parse, sheet, m_name, k_title,
                       dont_skip):
    print(filename, "readExcelAndGen")
    data_ignore_list, data_len, keys = read_keys(sheet, dont_skip)
    data_ignore_list = reversed(data_ignore_list)
    builder, mb, bs, ml, m, list_name = gen_process_begin(data_len, modules, dict_parse, m_name,
                                                          k_title)
    ext = "bytes"
    for i in reversed(range(2, sheet.max_row + 1)):
        data_ignore = next(data_ignore_list)
        if data_ignore:
            continue
        d = collections.OrderedDict()
        for k, v in keys.items():
            if isinstance(v[0], list):
                d[k] = (list(), v[1], v[2])
                for vv in v[0]:
                    d[k][0].append((sheet.cell(row=i, column=vv[0]), vv[1], vv[2]))
                trim_blank(d[k][0])
            else:
                d[k] = (sheet.cell(row=i, column=v[0]), v[1], v[2])
        gen_process(modules, d, builder, mb, bs, m, k_title, dict_parse, m_name)
    gen_process_end(builder, bs, ml, data_len, sheet.title, k_title, list_name, dir_map, ext)


def pre_gen(sheet_title, modules, dict_parse):
    pos = sheet_title.find("_")
    if pos < 0:
        m_name = sheet_title
    else:
        m_name = sheet_title[:pos]
    if m_name not in dict_parse:
        return False, None, None
    pos = sheet_title.find("_")
    m_name = sheet_title if pos < 0 else sheet_title[:pos]

    k_title = stringcase.capitalcase(m_name)
    if k_title not in modules:
        return False, None, None
    return True, m_name, k_title


def read_keys(sheet, dont_skip):
    row_extra = list(row[0].value for row in sheet.rows)
    rows = sheet.rows
    col_extra = list(t.value for t in next(rows))
    keys = None
    data_len = 0
    data_ignore_list = []
    for i in range(2, sheet.max_row + 1):
        row = next(rows)
        if row_extra[i - 1] == "skip" or row_extra[i - 1] == "server":
            data_ignore_list.append(True)
            continue
        row_is_empty = True
        for v in row:
            if v.value is not None:
                row_is_empty = False
                break
        if row_is_empty:
            data_ignore_list.append(True)
            continue
        if not keys:
            keys = collections.OrderedDict()
            for j in range(2, sheet.max_column + 1):
                if not dont_skip and (col_extra[j - 1] == "skip" or col_extra[j - 1] == "server"):
                    continue
                cell = sheet.cell(row=i, column=j)
                if cell.value is None:
                    continue
                cell_in_type_row = str(sheet.cell(row=i - 1, column=j).value)
                need_split = False
                need_kv = False
                is_list = False
                pattern = r"\[(.+)\]"
                r = re.search(pattern, cell_in_type_row)
                if r:
                    cell_in_type_row = r.group(1)
                    need_split = True
                pattern = r"\{(\w+)\}"
                r = re.search(pattern, cell_in_type_row)
                if r:
                    cell_in_type_row = r.group(1)
                    need_kv = True
                k = str(cell.value)
                pattern = r"(\w+)_(\d+)$"
                r = re.search(pattern, k)
                if r:
                    is_list = True
                    k = r.group(1)
                if is_list or need_split:
                    if k not in keys:
                        keys[k] = (list(), cell_in_type_row, need_kv)
                    elif not isinstance(keys[k][0], list):
                        keys.pop(k)
                        keys[k] = (list(), cell_in_type_row, need_kv)
                    keys[k][0].insert(int(r.group(2)) if r else 0, (j, need_split, need_kv))
                else:
                    if k not in keys:
                        keys[k] = (j, cell_in_type_row, need_kv)
            data_ignore_list.append(True)
            continue
        data_len += 1
        data_ignore_list.append(False)
    return data_ignore_list, data_len, keys


def gen_process_begin(data_len, modules, dict_parse, m_name, k_title):
    builder = flatbuffers.Builder(0)
    mb = dict(inspect.getmembers(builder, predicate=inspect.ismethod))
    bs = build_seq.BuildSeq()
    (list_name, _, _, _) = dict_parse[m_name][-1]
    ml = modules["{}List".format(k_title)].__dict__
    func_name = "%sListStart" % k_title
    bs.store_func(ml[func_name], [func_name, builder])
    bs.begin_sub_seq()
    func_name = "%sListStart%sVector" % (k_title, stringcase.capitalcase(list_name))
    bs.store_func(ml[func_name], [func_name, builder, data_len])
    m = modules[k_title].__dict__
    return builder, mb, bs, ml, m, list_name


def gen_process(modules, d, builder, mb, bs, m, k_title, dict_parse, m_name):
    func_name = "%sStart" % k_title
    bs.begin_sub_seq()
    bs.store_func(m[func_name], [func_name, builder])
    kv_index = 0
    for (name, is_vec, typ, _) in dict_parse[m_name][:-1]:
        d_key = name
        if name.endswith("K"):
            kv_index = 0
            d_key = name[:-1]
            typ = "uint"
        elif name.endswith("V"):
            kv_index = 1
            d_key = name[:-1]
        if d_key not in d:
            continue
        d_elem = d[d_key][0]
        name_title = stringcase.capitalcase(name)
        if is_vec:
            if not len(d_elem):
                continue
            func_name = "%sStart%sVector" % (k_title, name_title)
            bs.begin_sub_seq()
            # cal size
            size = 0
            for elem_v in reversed(d_elem):
                if elem_v[1]:
                    # split
                    src_str = str(elem_v[0].value)
                    str_arr = src_str.split("|")
                    size += len(str_arr)
                else:
                    size += 1
            bs.store_func(m[func_name], [func_name, builder, size])
            for elem_v in reversed(d_elem):
                if elem_v[1]:
                    if elem_v[0].value is None:
                        continue
                    # split
                    src_str = str(elem_v[0].value)
                    str_arr = src_str.split("|")
                    for elem_vvv in reversed(str_arr):
                        if elem_v[2]:
                            # kv
                            map_src_str = elem_vvv.split(":")
                            temp_value = format_str_value(map_src_str[kv_index], typ, False)
                        else:
                            temp_value = format_str_value(elem_vvv, typ, False)
                        store_list_func_and_arg(bs, mb, builder, temp_value, typ, modules)
                else:
                    temp_value = format_excel_value(elem_v[0], typ, False)
                    store_list_func_and_arg(bs, mb, builder, temp_value, typ, modules)
            bs.store_func(builder.EndVector, ["EndVector", size])
            sub_seq = bs.end_sub_seq()
            func_name = "%sAdd%s" % (k_title, name_title)
            bs.store_func(m[func_name], [func_name, builder, sub_seq])
        else:
            if d_elem.value is None:
                continue
            if d[d_key][2]:
                # kv
                src_str = str(d_elem.value)
                map_src_str = src_str.split(":")
                temp_value = format_str_value(map_src_str[kv_index], typ, False)
            else:
                temp_value = format_excel_value(d_elem, typ)
            if typ == "string":
                temp_value = builder.CreateString(temp_value or "")
            elif typ == "FP":
                ml = modules["FP"].__dict__
                temp_value = build_struct.BuildStruct(ml["Create{}".format(typ)],
                                                      [builder, int(temp_value * 4294967296)])
            func_name = "%sAdd%s" % (k_title, name_title)
            bs.store_func(m[func_name], [func_name, builder, temp_value])
    func_name = "%sEnd" % k_title
    bs.store_func(m[func_name], [func_name, builder])
    sub_seq = bs.end_sub_seq()
    bs.store_func(builder.PrependUOffsetTRelative, ["PrependUOffsetTRelative", sub_seq])


def gen_process_end(builder, bs, ml, data_len, k, k_title, list_name, dir_map, ext):
    bs.store_func(builder.EndVector, ["EndVector", data_len])
    sub_seq = bs.end_sub_seq()
    func_name = "%sListAdd%s" % (k_title, stringcase.capitalcase(list_name))
    bs.store_func(ml[func_name], [func_name, builder, sub_seq])
    func_name = "%sListEnd" % k_title
    bs.store_func(ml[func_name], [func_name, builder])
    fbs_data = bs.process()
    builder.Finish(fbs_data)
    buf = builder.Output()
    print(k, len(buf))
    with open(os.path.join(dir_map.out_bytes_dir, "{}.{}".format(k, ext)), "wb") as f:
        f.write(buf)
