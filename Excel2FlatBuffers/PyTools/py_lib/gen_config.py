import importlib
import os
import re
import sys
import threading

import openpyxl
import collections
import argparse
import stringcase
from openpyxl.reader import excel
from . import util
from .gen_config_process import gen_config_process


def read_keys(sheet, row_index):
    col_extra = list(t.value for t in next(sheet.rows))
    keys = collections.OrderedDict()
    for j in range(2, sheet.max_column + 1):
        if col_extra[j - 1] == "skip":
            continue
        cell = sheet.cell(row=row_index, column=j)
        if cell.value is None:
            continue
        cell_in_type_row = str(sheet.cell(row=row_index - 1, column=j).value)
        need_split = False
        need_kv = False
        is_list = False
        pattern = r"\[(.+)\]"
        r = re.search(pattern, cell_in_type_row)
        if r:
            cell_in_type_row = r.group(1)
            need_split = True
        pattern = r"\{(\w+)\}"
        r = re.search(pattern, cell_in_type_row)
        if r:
            cell_in_type_row = r.group(1)
            need_kv = True
        if cell_in_type_row not in ["bool", "int32", "sint32", "uint32", "string", "float", "FP"]:
            continue
        k = str(cell.value)
        pattern = r"(\w+)_(\d+)$"
        r = re.search(pattern, k)
        if r:
            is_list = True
            k = r.group(1)
        if is_list or need_split:
            if k not in keys:
                keys[k] = (list(), cell_in_type_row, need_kv)
            elif not isinstance(keys[k][0], list):
                keys.pop(k)
                keys[k] = (list(), cell_in_type_row, need_kv)
            keys[k][0].insert(int(r.group(2) if r else 0), (j, need_split, need_kv))
        else:
            if k not in keys:
                keys[k] = (j, cell_in_type_row, need_kv)
    return keys


def get_type_str_from_protobuf(type_name):
    if type_name == "bool":
        return "bool"
    if type_name == "sint32":
        return "int"
    if type_name == "int32" or type_name == "uint32":
        return "uint"
    if type_name == "string":
        return "string"
    if type_name == "float":
        return "float"
    if type_name == "FP":
        return "FP"
    raise "type not found: {}".format(type_name)


def read_excel(dir_map, file_path, sheet_name):
    book = openpyxl.reader.excel.load_workbook(file_path, data_only=True)
    if sheet_name not in book.sheetnames:
        print("{} ignoredd".format(sheet_name))
        book.close()
        return None
    print(sheet_name)
    # pos = fname.find("_")
    # if pos >= 0:
    #     fname = fname[:pos]
    sheet = book[sheet_name]
    if sheet.max_row < 1 or sheet.max_column < 1:
        print("{} sheet is None".format(sheet.title))
        book.close()
        return None
    rows = sheet.rows
    columns = sheet.columns
    # colExtra = list(t.value for t in next(rows))
    row_extra = list(t.value for t in next(columns))
    keys = None
    for i in range(2, sheet.max_row + 1):
        row = next(rows)
        if row_extra[i - 1] == "skip":
            continue
        row_is_empty = True
        for v in row:
            if v.value is not None:
                row_is_empty = False
                break
        if row_is_empty:
            continue
        if not keys:
            keys = read_keys(sheet, i)
            write_csv(dir_map, sheet)
            book.close()
            return keys
    print(sheet_name + "keys is null")
    book.close()
    return None


def write_fp_proto(dir_map):
    file_path = os.path.join(dir_map.out_proto_dir, "FP.fbs")
    with open(file_path, "w") as f:
        f.write("namespace {};\n\n".format(dir_map.fbs_namespace))
        f.write("struct FP {\n")
        f.write("    raw:long;\n")
        f.write("}\n\n")
        f.write("file_identifier \"CONF\";\n")
        f.write("file_extension \"game_config\";\n")
    return file_path


def write_proto(dir_map, key, first_data):
    pos = key.find("_")
    if pos >= 0:
        key = key[:pos]
    file_path = os.path.join(dir_map.out_proto_dir, "{}.fbs".format(key))
    if os.path.exists(file_path):
        return file_path, key
    print("write {} begin".format(key))
    with open(file_path, "w") as f:
        key_title = stringcase.capitalcase(key)
        # add import FP
        f.write("include \"FP.fbs\";\n\n")
        f.write("namespace {};\n\n".format(dir_map.fbs_namespace))
        f.write("table {} {{\n".format(key_title))
        for k, v in first_data.items():
            type_str = get_type_str_from_protobuf(v[1])
            if isinstance(v[0], list):
                if v[2]:
                    # kv
                    f.write("    {}K:[uint];\n".format(k))
                    f.write("    {}V:[{}];\n".format(k, type_str))
                else:
                    f.write("    {}:[{}];\n".format(k, type_str))
            else:
                if v[2]:
                    # kv
                    f.write("    {}K:uint;\n".format(k))
                    f.write("    {}V:{};\n".format(k, type_str))
                else:
                    f.write("    {}:{};\n".format(k, type_str))
        f.write("}\n\n")
        f.write("table {}List {{\n".format(key_title))
        f.write("    data:[{}];\n".format(key_title))
        f.write("}\n\n")
        f.write("file_identifier \"CONF\";\n")
        f.write("file_extension \"game_config\";\n")
    return file_path, key


def write_csv(dir_map, sheet):
    with open(os.path.join(dir_map.out_csv_dir, "{}.csv".format(sheet.title)), "wb") as f:
        skipColumn = set()
        for j in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=j)
            cellValue = format_csv_value(cell.value)
            if cellValue == "skip":
                skipColumn.add(j)
        
        needComma = False
        for j in range(2, sheet.max_column + 1):
            if j in skipColumn:
                continue
            cell = sheet.cell(row=4, column=j)
            cellValue = format_csv_value(cell.value)
            if cellValue == "":
                skipColumn.add(j)
            else:
                if not needComma:
                    needComma = True
                else:
                    f.write(b",")
                f.write(cellValue.encode("utf-8"))
                    
        f.write(b"\n")
        
        needComma = False
        for j in range(2, sheet.max_column + 1):
            if j in skipColumn:
                continue
            cell = sheet.cell(row=3, column=j)
            str_cell = format_csv_value(cell.value)
            if not needComma:
                needComma = True
            else:
                f.write(b",")
            if str_cell in ["bool", "int32", "sint32", "uint32"]:
                f.write(b"int")
            elif str_cell in ["float", "FP"]:
                f.write(b"float")
            else:
                f.write(b"string")
                
        f.write(b"\n")

        for i in range(5, sheet.max_row + 1):
            needComma = False
            cell = sheet.cell(row=i, column=1)
            cellValue = format_csv_value(cell)
            if cellValue == "skip":
                continue
            for j in range(2, sheet.max_column + 1):
                if j in skipColumn:
                    continue
                if not needComma:
                    needComma = True
                else:
                    f.write(b",")
                cell = sheet.cell(row=i, column=j)
                cellValue = format_csv_value(cell.value)
                cellValue.replace('"', '""')
                cellValue = '"' + cellValue + '"'
                f.write(cellValue.encode("utf-8"))
                
            f.write(b"\n")


def format_csv_value(value):
    if value is None:
        return ""
    return str(value)

#从ChooseExcel.txt文件中读取相应的Excel文件名
def read_txt_choose_excel():
    lines = []
    with open(os.path.join(os.getcwd(),"ChooseExcel.txt"), 'r') as file_to_read:
        while True:
            line = file_to_read.readline()
            if not line:
                break
            line = line.strip('\n')
            lines.append(line)
    if len(lines):
        return lines
    else:
        return None
    return None

def read_excel_and_process(dir_map):
    util.mk_out_dir(dir_map.out_proto_dir, True)
    util.mk_out_dir(dir_map.out_py_dir, True)
    util.mk_out_dir(dir_map.out_cs_dir, True)
    util.mk_out_dir(dir_map.out_cs_dir, True, [os.path.join(dir_map.out_cs_dir, "**/*.meta")])
    util.mk_out_dir(dir_map.out_lua_dir, True)
    util.mk_out_dir(dir_map.out_cpp_dir, True)
    util.mk_out_dir(dir_map.out_bytes_dir, True, [os.path.join(dir_map.out_bytes_dir, "**/*.meta")])
    util.mk_out_dir(dir_map.out_csv_dir, True)
    flatc_exec = os.path.join(dir_map.project_dir, "bin/flatc")
    if util.os_is_win32():
        flatc_exec = flatc_exec + ".exe"
    dict_parse = collections.OrderedDict()
    modules = collections.OrderedDict()

    fbs_file_path = write_fp_proto(dir_map)
    gen_config_process.flatc(dir_map, flatc_exec, fbs_file_path)
    module = read_proto(dir_map, "FP")
    modules["FP"] = module

    threads = []
    
    for dirpath, _, filenames in os.walk(dir_map.config_dir):
        filenames = filter(lambda x: re.match(r"^[^~]+\.xlsx$", x), filenames)
        for filename in filenames:
            fname, _ = os.path.splitext(filename)
            file_path = os.path.join(dirpath, filename)
            #读txt文件中对应的Excel文件
            if fname in read_txt_choose_excel():
                keys = read_excel(dir_map, file_path, fname)
            else:
                keys = None
            if keys is None:
                continue
            fbs_file_path, key = write_proto(dir_map, fname, keys)
            if fbs_file_path is None:
                continue
            if key in dict_parse:
                continue
            #启动线程
            t = threading.Thread(target=thread_process, args=(dict_parse, dir_map, flatc_exec, fbs_file_path, module, key , modules))
            threads.append(t)
            t.start()
    print("Loading...")
    for t in threads:
        t.join()
    print("All Over")
    gen_config_process.gen_bytes(dir_map, modules, dict_parse, False)

#线程写文件
def thread_process(dict_parse, dir_map, flatc_exec, fbs_file_path, module, key ,modules):
    name_list = gen_config_process.flatc(dir_map, flatc_exec, fbs_file_path)
    dict_parse[key] = name_list
    module, list_module = gen_config_process.read_proto(dir_map, key)
    key_title = stringcase.capitalcase(key)
    modules[key_title] = module
    modules["{}List".format(key_title)] = list_module

def read_proto(dir_map, key):
    rel_path = os.path.relpath(dir_map.out_py_dir, dir_map.project_dir).replace("\\", "/")
    lib_path = rel_path.replace("/", ".")
    rel_path = os.path.relpath(dir_map.out_py_dir, os.getcwd()).replace("\\", "/")
    key = stringcase.capitalcase(key)

    sys.path.append(os.path.abspath(util.get_current_path()))
    lib_util = importlib.util

    spec = lib_util.spec_from_file_location("{}.{}.{}".format(lib_path, dir_map.fbs_namespace, key),
                                            "{}/{}/{}.py".format(rel_path, dir_map.fbs_namespace, key))
    module = lib_util.module_from_spec(spec)
    spec.loader.exec_module(module)

    sys.path.pop()
    return module


def parse_arg(parser: argparse.ArgumentParser):
    parser.add_argument("--config-dir", dest="config_dir", type=str, help="config-dir")
    parser.add_argument("--out-proto-dir", dest="out_proto_dir", type=str, help="out-proto-dir")
    parser.add_argument("--out-cs-dir", dest="out_cs_dir", type=str, help="out-cs-dir")
    parser.add_argument("--out-lua-dir", dest="out_lua_dir", type=str, help="out-lua-dir")
    parser.add_argument("--out-cpp-dir", dest="out_cpp_dir", type=str, help="out-cpp-dir")
    parser.add_argument("--out-bytes-dir", dest="out_bytes_dir", type=str, help="out-bytes-dir")
    parser.add_argument("--out-csv-dir", dest="out_csv_dir", type=str, help="out-csv-dir")


# noinspection PyUnusedLocal
def main(args: argparse.Namespace):
    # gen_config_all()
    class DirMap:
        fbs_namespace = "GameConfig"
        cur_dir = None
        project_dir = None
        config_dir = None
        out_proto_dir = None
        out_py_dir = None
        out_cs_dir = None
        out_lua_dir = None
        out_cpp_dir = None
        out_bytes_dir = None
        out_csv_dir = None

    dir_map = DirMap()
    dir_map.cur_dir = util.get_current_path()                                                             #绝对路径D:\...\Excel2FlatBuffers\Pytools\py_lib
    dir_map.project_dir = os.path.abspath(os.path.join(dir_map.cur_dir, os.path.pardir))                  #绝对路径D:\...\Excel2FlatBuffers\Pytools
    dir_map.config_dir = args.config_dir or os.path.join(dir_map.project_dir, "config")                   #ExcelTable
    dir_map.out_proto_dir = args.out_proto_dir or os.path.join(dir_map.project_dir, "out/proto")          #out\proto
    dir_map.out_py_dir = os.path.join(dir_map.project_dir, "py_lib/out/py")                               #绝对路径D:\...\Excel2FlatBuffers\Pytools\py_lib/out/py
    dir_map.out_cs_dir = args.out_cs_dir or os.path.join(dir_map.project_dir, "out/cs")                   #out\cs
    dir_map.out_lua_dir = args.out_lua_dir or os.path.join(dir_map.project_dir, "out/lua")                #out\lua
    dir_map.out_cpp_dir = args.out_cpp_dir or os.path.join(dir_map.project_dir, "out/cpp")                #out\cpp
    dir_map.out_bytes_dir = args.out_bytes_dir or os.path.join(dir_map.project_dir, "out/bytes")          #out\bytes 
    dir_map.out_csv_dir = args.out_csv_dir or os.path.join(dir_map.project_dir, "out/csv")                #out\csv

    read_excel_and_process(dir_map)
